from Mistika.Qt import QColor
from Mistika.classes import Cconnector
from Mistika.classes import CuniversalPath, CnameConvention
import xml.etree.ElementTree as ET
import openpyxl
import os


def init(self):
    self.color=QColor(0x677688)
# creating properties
    self.addProperty("configXML")
    self.addProperty("dstFolder")
#creating connectors
    self.addConnector("xml",Cconnector.CONNECTOR_TYPE_INPUT,Cconnector.MODE_REQUIRED)
    self.addConnector("xlsx",Cconnector.CONNECTOR_TYPE_OUTPUT,Cconnector.MODE_OPTIONAL)
    self.setAcceptConnectors(True,"files")
#configuring the node
    self.bypassSupported=True
    self.color=QColor(0,180,180)
    return True


def isReady(self):
    if self.bypassSupported and self.bypassEnabled:
        return True
    res=True
    # if not self.dstFolder:
    #     res=self.critical("XMLtoXLSX:dstFolder:empty","destination folder path can not be empty") and res
    if not self.configXML:
        res=self.critical("XMLtoXLSX:configXML:empty","config file path can not be empty") and res
    if len(self.dstFolder)>0 and not self.dstFolder.endswith("/"):
        res=self.critical("XMLtoXLSX:dstFolder:naming","destination folder path must end with '/' character") and res
    if len(self.dstFolder)>0 and not os.path.isdir(self.dstFolder):
        res=self.critical("XMLtoXLSX:dstFolderL:notFound","destination folder not found:'{}'".format(self.dstFolder)) and res
    if len(self.configXML)>0 and not os.path.isfile(self.configXML):
        res=self.critical("XMLtoXLSX:configXML:notFound","config XML file not found:'{}'".format(self.configXML)) and res
    return res


def process(self):
    #And finally, Do the thing here!
    res=True
    inputs=self.getConnectorsByType(Cconnector.CONNECTOR_TYPE_INPUT)
    output=self.getFirstConnectorByName("xlsx")    


    output.clearUniversalPaths()
    if self.bypassEnabled:
        return True


    for c in inputs:                                 
        for up in c.getUniversalPaths():
            if self.isCancelled():
                return False
            files = up.getAllFiles()
            for f in files:
                if not f.endswith(".xml"):
                    self.warning("XMLtoXLSX:notXML","The file'{}' is not an XML, can not convert to XLSX".format(f))
                    continue
       
                # Load the configuration XML document using the ElementTree library
                try:
                    config = ET.parse(self.configXML)
                    rootConfig = config.getroot()
                except ET.ParseError as e:
                    return self.critical("XMLtoXLSX:parseError","Error parsing configXML file '{}': {}".format(self.configXML, e))
                
                # Load the XML document using the ElementTree library
                try:
                    xml = ET.parse(f, parser=ET.XMLParser(encoding="utf-8"))
                    rootXml = xml.getroot()
                    if "itemRoot" in rootConfig.attrib:
                        rootXml = xml.find(rootConfig.attrib["itemRoot"])
                except ET.ParseError as e:
                    res = self.critical("XMLtoXLSX:parseError","Error parsing data XML file '{}': {}".format(f, e)) and res
                    continue

                # Create a new XLSX workbook
                wb = openpyxl.Workbook()
                ws = wb.active

                #Write the data from the XML document to the XLSX file
                for i, item in enumerate(rootConfig):
                    
                    #Write column titles
                    if "to" in item.attrib:
                        titleValue = item.attrib["to"]
                    else:
                        try:
                            titleArray = item.attrib["from"]
                        except KeyError:
                            return self.critical("XMLtoXLSX:parseError","Error: config xml must follow strict rules, see documentation to learn more about it")
                        titleSplit = titleArray.split("/")
                        titleValue = titleSplit[len(titleSplit)-1]
                    ws.cell(row = 1, column = i+1, value = titleValue)
                    
                    for j, child in enumerate(rootXml):
                        if item.attrib["from"] == rootXml.tag:
                            cellValue = rootXml
                        elif item.attrib["from"] == child.tag:
                            cellValue = child
                        else:
                            search = item.attrib["from"]
                            if search.startswith(child.tag):
                                search = search.replace(child.tag + "/", "")

                            #If it has extra attributes, use them to filter. Create proper XPath xpresion
                            for attrib in item.attrib:
                                if attrib != "from" and attrib != "to" and attrib != "getValueFromProperty":
                                    search = search + "[@" + attrib + "='" + item.attrib[attrib] + "']"
                            try:
                                #print(search)
                                cellValue = child.find(search)
                            except IndexError:
                                cellValue = None


                        if not cellValue == None:
                            if "getValueFromProperty" in item.attrib:
                                propertyName = item.attrib["getValueFromProperty"]
                                cellValue = cellValue.attrib[propertyName]
                            else:
                                cellValue = cellValue.text
                        else:
                            if item.attrib["from"] == "python":                     
                                try: 
                                    print(item.attrib["to"])
                                    exec(item.text, globals(), locals())
                                    func = locals()["xml2xlsxPythonProperty"]
                                    cellValue = func(self, child)
                                except Exception as e:
                                    print("Error evaluating function definition:", e)
                                locals().pop("xml2xlsxPythonProperty", None)
                            else:
                                cellValue = ""


                        ws.cell(row = j+2, column = i+1, value = cellValue)
                # Save the XLSX file
                try:
                    if len(self.dstFolder)==0:
                        finalPath = self.dstFolder + os.path.splitext(f)[0] + ".xlsx"
                    else:
                        finalPath = self.dstFolder + os.path.splitext(os.path.basename(f))[0] + ".xlsx"


                    wb.save(finalPath)
                    nc=CnameConvention("[path][baseName][.ext]")
                    up = CuniversalPath(nc, finalPath)
                    output.addUniversalPath(up)
                except PermissionError:
                    return self.critical("XMLtoXLSX:configXML:noPermission","No permission to acces:'{}'. Make sure the file is not opened by another program".format(finalPath))
    
    return res